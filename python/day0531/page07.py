from openpyxl import *
import datetime

wb = Workbook()
ws = wb.active
ws['A1'] = 44
ws["A2"] = "korea"
ws["A3"] = "한글"

ws["C5"] = datetime.datetime.now()


ws.append([4,5,6])
wb.save("sam.xlsx")
wb.close()
print("end")
